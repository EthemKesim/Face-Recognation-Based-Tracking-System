from __future__ import annotations

import json
import re
import csv
import io
import sqlite3
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


PROJECT_SOURCE_DIR = Path(__file__).resolve().parents[2]
if str(PROJECT_SOURCE_DIR) not in sys.path:
    sys.path.insert(0, str(PROJECT_SOURCE_DIR))

from time_override import get_current_datetime

DB_PATH = PROJECT_SOURCE_DIR / "face_records.db"
LOG_PATH = PROJECT_SOURCE_DIR / "attendance_logs.txt"
MAIN_SCRIPT_PATH = PROJECT_SOURCE_DIR / "main_recognition.py"

LOG_PATTERN = re.compile(r"^(?P<timestamp>\d{2}/\d{2}/\d{4} \d{2}:\d{2}:\d{2}) - (?P<name>.+?) - (?P<status>.+)$")
LOG_DATETIME_FORMAT = "%d/%m/%Y %H:%M:%S"
API_DATE_FORMAT = "%Y-%m-%d"


@dataclass
class AttendanceRecord:
    employee_id: int | None
    employee_name: str
    work_date: str
    entry_time: str | None
    exit_time: str | None
    current_status: str
    last_event_type: str
    notes: list[str]
    events: list[dict[str, Any]]

    def to_dict(self) -> dict[str, Any]:
        return {
            "employee_id": self.employee_id,
            "employee_name": self.employee_name,
            "date": self.work_date,
            "entry_time": self.entry_time,
            "exit_time": self.exit_time,
            "current_status": self.current_status,
            "event_type": self.last_event_type,
            "notes": self.notes,
            "events": self.events,
        }


def load_registered_users() -> list[dict[str, Any]]:
    if not DB_PATH.exists():
        return []

    with sqlite3.connect(DB_PATH) as connection:
        cursor = connection.cursor()
        if table_exists(cursor, "employees"):
            cursor.execute(
                """
                SELECT id, full_name, status, photo_path
                , department_role
                FROM employees
                ORDER BY full_name COLLATE NOCASE, id
                """
            )
            rows = cursor.fetchall()
            return [
                {
                    "id": row[0],
                    "name": row[1],
                    "status": row[2],
                    "photo_path": row[3],
                    "department_role": row[4],
                    "image_url": ("/" + row[3].replace("\\", "/").lstrip("/")) if row[3] else None,
                    "face_registered": row[2] == "active",
                }
                for row in rows
            ]

        cursor.execute("SELECT id, name FROM users ORDER BY name COLLATE NOCASE, id")
        rows = cursor.fetchall()

    return [{"id": row[0], "name": row[1], "status": "active", "photo_path": None, "department_role": None, "image_url": None, "face_registered": True} for row in rows]


def table_exists(cursor: sqlite3.Cursor, table_name: str) -> bool:
    cursor.execute("SELECT 1 FROM sqlite_master WHERE type = 'table' AND name = ?", (table_name,))
    return cursor.fetchone() is not None


def load_structured_attendance_records() -> list[AttendanceRecord]:
    if not DB_PATH.exists():
        return []

    with sqlite3.connect(DB_PATH) as connection:
        cursor = connection.cursor()
        if not table_exists(cursor, "attendance_logs") or not table_exists(cursor, "employees"):
            return []

        cursor.execute(
            """
            SELECT
                attendance_logs.employee_id,
                employees.full_name,
                attendance_logs.date,
                attendance_logs.entry_time,
                attendance_logs.exit_time,
                attendance_logs.attendance_status
            FROM attendance_logs
            INNER JOIN employees ON employees.id = attendance_logs.employee_id
            ORDER BY attendance_logs.date DESC, employees.full_name COLLATE NOCASE
            """
        )
        rows = cursor.fetchall()

    records: list[AttendanceRecord] = []
    for row in rows:
        notes = build_structured_attendance_notes(row[5], row[3], row[4])
        current_status = derive_structured_current_status(row[3], row[4], row[5])
        last_event_type = "CHECK-OUT" if row[4] else "CHECK-IN" if row[3] else "UNKNOWN"
        events = build_structured_event_list(row[1], row[2], row[3], row[4], row[5])
        records.append(
            AttendanceRecord(
                employee_id=row[0],
                employee_name=row[1],
                work_date=row[2],
                entry_time=row[3],
                exit_time=row[4],
                current_status=current_status,
                last_event_type=last_event_type,
                notes=notes,
                events=events,
            )
        )

    return records


def parse_log_events() -> list[dict[str, Any]]:
    # Log parsing happens here. Each line is interpreted from the current
    # `attendance_logs.txt` format produced by the Python recognition script.
    if not LOG_PATH.exists():
        return []

    events: list[dict[str, Any]] = []
    for line_number, raw_line in enumerate(LOG_PATH.read_text(encoding="utf-8").splitlines(), start=1):
        line = raw_line.strip()
        if not line:
            continue

        match = LOG_PATTERN.match(line)
        if not match:
            events.append(
                {
                    "id": f"log-{line_number}",
                    "line_number": line_number,
                    "raw": line,
                    "timestamp": None,
                    "date": None,
                    "time": None,
                    "employee_name": None,
                    "status": "UNPARSED",
                    "event_type": "UNKNOWN",
                    "status_group": "unknown",
                    "notes": ["This line could not be parsed by the dashboard."],
                }
            )
            continue

        timestamp = datetime.strptime(match.group("timestamp"), LOG_DATETIME_FORMAT)
        name = match.group("name").strip()
        status = match.group("status").strip()

        events.append(
            {
                "id": f"log-{line_number}",
                "line_number": line_number,
                "raw": line,
                "timestamp": timestamp.isoformat(),
                "date": timestamp.strftime(API_DATE_FORMAT),
                "time": timestamp.strftime("%H:%M:%S"),
                "employee_name": name,
                "status": status,
                "event_type": determine_event_type(status),
                "status_group": determine_status_group(status),
                "notes": build_notes(status),
                "_dt": timestamp,
            }
        )

    events.sort(key=lambda event: event.get("_dt") or datetime.min, reverse=True)
    return events


def determine_event_type(status: str) -> str:
    if status.startswith("OVERTIME"):
        return "CHECK-OUT"
    if status.startswith("CHECK-IN") or status.startswith("WARNING: Late") or status.startswith("VIOLATION: Late"):
        return "CHECK-IN"
    if status.startswith("CHECK-OUT"):
        return "CHECK-OUT"
    return "UNKNOWN"


def determine_status_group(status: str) -> str:
    normalized = status.upper()
    if "VIOLATION" in normalized:
        return "violation"
    if "WARNING" in normalized:
        return "warning"
    if "LATE" in normalized:
        return "late"
    if "LUNCH BREAK" in normalized:
        return "lunch"
    if "OVERTIME" in normalized:
        return "overtime"
    if normalized.startswith("CHECK-OUT"):
        return "checkout"
    if normalized.startswith("CHECK-IN"):
        return "checkin"
    return "neutral"


def build_notes(status: str) -> list[str]:
    notes: list[str] = []
    if "Lunch Break" in status:
        notes.append("Lunch break window")
    if "WARNING: Late" in status:
        notes.append("Late warning")
    if "VIOLATION: Late" in status:
        notes.append("Late violation")
    if "OVERTIME" in status:
        notes.append("Overtime detected")
    if "After 18:00" in status:
        notes.append("Checked out after overtime threshold")
    return notes


def build_attendance_records(
    events: list[dict[str, Any]],
    users: list[dict[str, Any]],
) -> list[AttendanceRecord]:
    user_map = {user["name"]: user["id"] for user in users}
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)

    for event in events:
        if not event.get("employee_name") or not event.get("date"):
            continue
        grouped[(event["employee_name"], event["date"])].append(event)

    records: list[AttendanceRecord] = []
    for (employee_name, work_date), employee_events in grouped.items():
        ordered_events = sorted(employee_events, key=lambda item: item["_dt"])
        entry_time = next(
            (event["time"] for event in ordered_events if event["event_type"] == "CHECK-IN"),
            None,
        )
        exit_time = next(
            (event["time"] for event in reversed(ordered_events) if event["event_type"] == "CHECK-OUT"),
            None,
        )
        last_event = ordered_events[-1]
        current_status = derive_current_status(ordered_events)

        records.append(
            AttendanceRecord(
                employee_id=user_map.get(employee_name),
                employee_name=employee_name,
                work_date=work_date,
                entry_time=entry_time,
                exit_time=exit_time,
                current_status=current_status,
                last_event_type=last_event["event_type"],
                notes=collect_record_notes(ordered_events, current_status),
                events=[serialize_event(event) for event in reversed(ordered_events)],
            )
        )

    records.sort(key=lambda record: (record.work_date, record.employee_name), reverse=True)
    return records


def serialize_event(event: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": event["id"],
        "timestamp": event["timestamp"],
        "date": event["date"],
        "time": event["time"],
        "employee_name": event["employee_name"],
        "employee_image_url": event.get("employee_image_url"),
        "status": event["status"],
        "event_type": event["event_type"],
        "status_group": event["status_group"],
        "notes": event["notes"],
        "raw": event["raw"],
    }


def derive_current_status(events: list[dict[str, Any]]) -> str:
    last_event = events[-1]
    status = last_event["status"]
    if last_event["event_type"] == "CHECK-IN":
        if "Lunch Break" in status:
            return "Lunch Break"
        return "Still Inside"
    if last_event["event_type"] == "CHECK-OUT":
        if "OVERTIME" in status:
            return "Checked Out (Overtime)"
        return "Checked Out"
    return status


def collect_record_notes(events: list[dict[str, Any]], current_status: str) -> list[str]:
    notes: list[str] = []
    if current_status == "Still Inside":
        notes.append("No check-out event recorded yet for this day.")
    if any("WARNING: Late" in event["status"] for event in events):
        notes.append("Includes a late warning event.")
    if any("VIOLATION: Late" in event["status"] for event in events):
        notes.append("Includes a late violation event.")
    if any("Lunch Break" in event["status"] for event in events):
        notes.append("Contains a lunch break state.")
    if any("OVERTIME" in event["status"] for event in events):
        notes.append("Contains overtime.")
    return notes


def build_structured_attendance_notes(
    attendance_status: str,
    entry_time: str | None,
    exit_time: str | None,
) -> list[str]:
    notes: list[str] = []
    if not exit_time and entry_time:
        notes.append("No check-out event recorded yet for this day.")
    if attendance_status == "late":
        notes.append("Includes a late check-in.")
    if attendance_status == "absent":
        notes.append("Marked absent.")
    return notes


def derive_structured_current_status(
    entry_time: str | None,
    exit_time: str | None,
    attendance_status: str,
) -> str:
    if attendance_status == "absent":
        return "Absent"
    if exit_time:
        return "Checked Out"
    if entry_time:
        if attendance_status == "late":
            return "Late Check-In"
        return "Still Inside"
    return "No activity"


def build_structured_event_list(
    employee_name: str,
    work_date: str,
    entry_time: str | None,
    exit_time: str | None,
    attendance_status: str,
) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    if exit_time:
        events.append(
            {
                "id": f"db-{employee_name}-{work_date}-exit",
                "timestamp": f"{work_date}T{exit_time}",
                "date": work_date,
                "time": exit_time,
                "employee_name": employee_name,
                "status": "CHECK-OUT",
                "event_type": "CHECK-OUT",
                "status_group": "checkout",
                "notes": [],
                "raw": "Structured database record",
            }
        )
    if entry_time:
        entry_status = "WARNING: Late" if attendance_status == "late" else "CHECK-IN"
        events.append(
            {
                "id": f"db-{employee_name}-{work_date}-entry",
                "timestamp": f"{work_date}T{entry_time}",
                "date": work_date,
                "time": entry_time,
                "employee_name": employee_name,
                "status": entry_status,
                "event_type": "CHECK-IN",
                "status_group": "warning" if attendance_status == "late" else "checkin",
                "notes": build_structured_attendance_notes(attendance_status, entry_time, exit_time),
                "raw": "Structured database record",
            }
        )
    return events


def get_dashboard_data() -> dict[str, Any]:
    users = load_registered_users()
    user_image_map: dict[str, str | None] = {u["name"]: u.get("image_url") for u in users}

    events = parse_log_events()
    for event in events:
        name = event.get("employee_name")
        event["employee_image_url"] = user_image_map.get(name) if name else None

    records = load_structured_attendance_records() or build_attendance_records(events, users)
    today = get_current_datetime().date().strftime(API_DATE_FORMAT)
    todays_records = [record for record in records if record.work_date == today]
    latest_detection = serialize_event(events[0]) if events else None

    def enrich_record(record: AttendanceRecord) -> dict[str, Any]:
        d = record.to_dict()
        d["employee_image_url"] = user_image_map.get(record.employee_name)
        return d

    return {
        "users": users,
        "events": [serialize_event(event) for event in events],
        "records": [enrich_record(r) for r in records],
        "today_records": [enrich_record(r) for r in todays_records],
        "latest_detection": latest_detection,
        "summary": build_summary(users, todays_records, events),
    }


def build_summary(
    users: list[dict[str, Any]],
    todays_records: list[AttendanceRecord],
    events: list[dict[str, Any]],
) -> dict[str, Any]:
    present_today = [record for record in todays_records if record.entry_time]
    late_today = [
        record
        for record in todays_records
        if any("Late" in event["status"] for event in record.events)
    ]
    checked_out_today = [record for record in todays_records if record.exit_time]
    overtime_today = [
        record
        for record in todays_records
        if any("OVERTIME" in event["status"] for event in record.events)
    ]

    active_users = [u for u in users if u.get("status") != "inactive"]
    present_names = {record.employee_name for record in present_today}
    absent_employees = [
        {"id": u["id"], "name": u["name"], "image_url": u.get("image_url")}
        for u in active_users
        if u["name"] not in present_names
    ]

    return {
        "total_registered_employees": len(users),
        "present_today": len(present_today),
        "late_today": len(late_today),
        "checked_out_today": len(checked_out_today),
        "overtime_employees": len(overtime_today),
        "absent_today": len(absent_employees),
        "absent_employees": absent_employees,
        "admin_alerts": build_admin_alerts(events),
        "recent_detections": [serialize_event(event) for event in events[:8]],
    }


def build_admin_alerts(
    events: list[dict[str, Any]],
    late_threshold: int = 3,
    window_days: int = 7,
) -> list[dict[str, Any]]:
    today = get_current_datetime().date()
    window_start = today - timedelta(days=window_days - 1)
    late_by_employee: dict[str, dict[str, Any]] = {}

    for event in events:
        employee_name = event.get("employee_name")
        event_date = event.get("date")
        if not employee_name or not event_date:
            continue
        if event.get("status_group") not in {"late", "warning", "violation"}:
            continue

        try:
            parsed_date = datetime.strptime(event_date, API_DATE_FORMAT).date()
        except ValueError:
            continue

        if parsed_date < window_start or parsed_date > today:
            continue

        alert = late_by_employee.setdefault(
            employee_name,
            {
                "employee_name": employee_name,
                "employee_image_url": event.get("employee_image_url"),
                "late_dates": set(),
                "latest_timestamp": event.get("timestamp"),
                "latest_status": event.get("status"),
            },
        )
        alert["late_dates"].add(event_date)
        if event.get("_dt") and event.get("timestamp"):
            latest_dt = datetime.fromisoformat(alert["latest_timestamp"]) if alert.get("latest_timestamp") else datetime.min
            if event["_dt"] > latest_dt:
                alert["latest_timestamp"] = event.get("timestamp")
                alert["latest_status"] = event.get("status")

    alerts = []
    for alert in late_by_employee.values():
        late_dates = sorted(alert["late_dates"], reverse=True)
        if len(late_dates) < late_threshold:
            continue
        alerts.append(
            {
                "type": "repeated_late",
                "severity": "warning",
                "employee_name": alert["employee_name"],
                "employee_image_url": alert.get("employee_image_url"),
                "count": len(late_dates),
                "threshold": late_threshold,
                "window_days": window_days,
                "late_dates": late_dates,
                "latest_timestamp": alert.get("latest_timestamp"),
                "latest_status": alert.get("latest_status"),
                "message": f"{alert['employee_name']} was late {len(late_dates)} times in the last {window_days} days.",
            }
        )

    alerts.sort(key=lambda item: (item["count"], item.get("latest_timestamp") or ""), reverse=True)
    return alerts[:6]


def filter_events(
    events: list[dict[str, Any]],
    name_query: str | None = None,
    work_date: str | None = None,
    status_filter: str | None = None,
) -> list[dict[str, Any]]:
    filtered = []
    name_query_normalized = (name_query or "").strip().lower()
    status_filter_normalized = (status_filter or "").strip().lower()

    for event in events:
        if name_query_normalized and name_query_normalized not in (event.get("employee_name") or "").lower():
            continue
        if work_date and event.get("date") != work_date:
            continue
        if status_filter_normalized:
            haystack = f'{event.get("status", "")} {event.get("status_group", "")} {event.get("event_type", "")}'.lower()
            if status_filter_normalized not in haystack:
                continue
        filtered.append(serialize_event(event))

    return filtered


def filter_records(
    records: list[dict[str, Any]],
    name_query: str | None = None,
    work_date: str | None = None,
    status_filter: str | None = None,
) -> list[dict[str, Any]]:
    filtered = []
    name_query_normalized = (name_query or "").strip().lower()
    status_filter_normalized = (status_filter or "").strip().lower()

    for record in records:
        if name_query_normalized and name_query_normalized not in record["employee_name"].lower():
            continue
        if work_date and record["date"] != work_date:
            continue
        if status_filter_normalized:
            record_haystack = " ".join(
                [record["current_status"], record["event_type"], " ".join(record["notes"])]
            ).lower()
            if status_filter_normalized not in record_haystack:
                continue
        filtered.append(record)

    return filtered


def build_employee_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    users = data["users"]
    events = data["events"]
    records = data["records"]
    latest_event_by_name: dict[str, dict[str, Any]] = {}
    todays_record_by_name: dict[str, dict[str, Any]] = {}
    today = get_current_datetime().date().strftime(API_DATE_FORMAT)

    for event in events:
        name = event.get("employee_name")
        if name and name not in latest_event_by_name:
            latest_event_by_name[name] = event

    for record in records:
        if record["date"] == today and record["employee_name"] not in todays_record_by_name:
            todays_record_by_name[record["employee_name"]] = record

    rows = []
    for user in users:
        latest_event = latest_event_by_name.get(user["name"])
        today_record = todays_record_by_name.get(user["name"])
        rows.append(
            {
                "id": user["id"],
                "name": user["name"],
                "status": user.get("status", "active"),
                "image_url": user.get("image_url"),
                "department_role": user.get("department_role"),
                "face_registered": user["face_registered"],
                "last_seen": latest_event["timestamp"] if latest_event else None,
                "current_status": today_record["current_status"] if today_record else "Absent / No activity today",
                "today_record": today_record,
            }
        )

    return rows


def get_employee_detail(employee_id: int, data: dict[str, Any]) -> dict[str, Any] | None:
    employee = next((user for user in data["users"] if user["id"] == employee_id), None)
    if not employee:
        return None

    employee_records = [record for record in data["records"] if record["employee_id"] == employee_id]
    employee_events = [event for event in data["events"] if event.get("employee_name") == employee["name"]]
    latest_event = employee_events[0] if employee_events else None

    late_count = sum(
        1 for event in employee_events if "WARNING: Late" in event["status"] or "VIOLATION: Late" in event["status"]
    )
    overtime_count = sum(1 for event in employee_events if "OVERTIME" in event["status"])

    return {
        "id": employee["id"],
        "name": employee["name"],
        "image_url": employee.get("image_url"),
        "department_role": employee.get("department_role"),
        "face_registered": employee["face_registered"],
        "latest_attendance_state": employee_records[0]["current_status"] if employee_records else "No attendance records",
        "latest_event": latest_event,
        "history": employee_records,
        "late_history_count": late_count,
        "overtime_history_count": overtime_count,
    }


def get_status_rules() -> dict[str, Any]:
    return {
        "source": str(MAIN_SCRIPT_PATH),
        "database_path": str(DB_PATH),
        "log_path": str(LOG_PATH),
        "rules": [
            {"name": "Morning Warning", "time": "09:15", "description": "CHECK-IN after 09:15 becomes WARNING: Late (Morning)."},
            {"name": "Morning Violation", "time": "09:30", "description": "CHECK-IN after 09:30 becomes VIOLATION: Late (Morning)."},
            {"name": "Lunch Break Window", "time": "12:00 - 13:15", "description": "Any check-in or check-out inside this range includes the Lunch Break label."},
            {"name": "Afternoon Warning", "time": "13:30", "description": "Post-lunch CHECK-IN after 13:30 becomes WARNING: Late (Afternoon)."},
            {"name": "Afternoon Violation", "time": "13:45", "description": "Post-lunch CHECK-IN after 13:45 becomes VIOLATION: Late (Afternoon)."},
            {"name": "Overtime Threshold", "time": "18:00", "description": "CHECK-OUT after 18:00 is labeled CHECK-OUT (After 18:00)."},
            {"name": "Work Duration Overtime", "time": "9 hours worked", "description": "If time between check-in and check-out exceeds 9 hours, the system writes OVERTIME: <hours>."},
        ],
    }


def json_bytes(payload: Any) -> bytes:
    return json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")



EXPORT_CSV_HEADERS = [
    "Employee ID",
    "Employee Name",
    "Date",
    "Entry Time",
    "Exit Time",
    "Current Status",
    "Event Type",
    "Notes",
]


def attendance_records_to_csv(records: list[dict[str, Any]]) -> bytes:
    """Convert a list of attendance record dicts into CSV bytes (UTF-8 BOM)."""
    buffer = io.StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPORT_CSV_HEADERS)

    for record in records:
        notes = record.get("notes") or []
        notes_text = "; ".join(notes) if isinstance(notes, list) else str(notes)
        writer.writerow(
            [
                record.get("employee_id") if record.get("employee_id") is not None else "",
                record.get("employee_name", ""),
                record.get("date", ""),
                record.get("entry_time") or "",
                record.get("exit_time") or "",
                record.get("current_status", ""),
                record.get("event_type", ""),
                notes_text,
            ]
        )

    return "\ufeff".encode("utf-8") + buffer.getvalue().encode("utf-8")


def build_export_filename(work_date: str | None = None) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if work_date:
        return f"attendance_{work_date}_{timestamp}.csv"
    return f"attendance_export_{timestamp}.csv"


def build_xlsx_filename(report_type: str, work_date: str | None = None) -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if report_type and report_type != "all":
        return f"attendance_{report_type}_{timestamp}.xlsx"
    if work_date:
        return f"attendance_{work_date}_{timestamp}.xlsx"
    return f"attendance_export_{timestamp}.xlsx"


XLSX_HEADERS = [
    "Employee ID", "Employee Name", "Date", "Entry Time", "Exit Time",
    "Duration", "Current Status", "Event Type", "Notes",
]


def _calc_duration(entry: str | None, exit_: str | None) -> str:
    if not entry or not exit_:
        return ""
    try:
        def _secs(t: str) -> int:
            parts = t.split(":")
            h, m = int(parts[0]), int(parts[1])
            s = int(parts[2]) if len(parts) > 2 else 0
            return h * 3600 + m * 60 + s
        diff = _secs(exit_) - _secs(entry)
        if diff <= 0:
            return ""
        hours, rem = divmod(diff, 3600)
        mins = rem // 60
        return f"{hours}h {mins}m" if hours > 0 else f"{mins}m"
    except (ValueError, IndexError):
        return ""


def attendance_records_to_xlsx(records: list[dict[str, Any]]) -> bytes | None:
    if not OPENPYXL_AVAILABLE:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for record in records:
        notes = record.get("notes") or []
        notes_text = "; ".join(notes) if isinstance(notes, list) else str(notes)
        entry = record.get("entry_time") or ""
        exit_ = record.get("exit_time") or ""
        ws.append([
            record.get("employee_id") if record.get("employee_id") is not None else "",
            record.get("employee_name", ""),
            record.get("date", ""),
            entry,
            exit_,
            _calc_duration(entry or None, exit_ or None),
            record.get("current_status", ""),
            record.get("event_type", ""),
            notes_text,
        ])

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def build_report_records(
    report_type: str,
    all_records: list[dict[str, Any]],
    users: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    today = get_current_datetime().date()
    today_str = today.strftime(API_DATE_FORMAT)

    if report_type == "daily":
        return [r for r in all_records if r.get("date") == today_str]

    if report_type == "weekly":
        week_start = (today - timedelta(days=6)).strftime(API_DATE_FORMAT)
        return [r for r in all_records if week_start <= (r.get("date") or "") <= today_str]

    if report_type == "monthly":
        month_prefix = today.strftime("%Y-%m")
        return [r for r in all_records if (r.get("date") or "").startswith(month_prefix)]

    if report_type == "late":
        return [
            r for r in all_records
            if "late" in (r.get("current_status") or "").lower()
            or any("late" in (n or "").lower() for n in (r.get("notes") or []))
        ]

    if report_type == "absent":
        active_users = [u for u in users if u.get("status") != "inactive"]
        today_names = {r["employee_name"] for r in all_records if r.get("date") == today_str}
        return [
            {
                "employee_id": u["id"],
                "employee_name": u["name"],
                "date": today_str,
                "entry_time": None,
                "exit_time": None,
                "current_status": "Absent",
                "event_type": "ABSENT",
                "notes": [],
            }
            for u in active_users
            if u["name"] not in today_names
        ]

    return all_records
